#!/usr/bin/env python3
"""Document Intelligence Pipeline — OCR extraction, enrichment, template filling.

Flow: file → OCR(doubao vision) → structured JSON → context graph enrichment → template fill
"""

import argparse, base64, json, os, subprocess, sys
from datetime import datetime
from pathlib import Path

import httpx

# ── Config ──────────────────────────────────────────────────
LITELLM_URL = os.environ.get("LITELLM_BASE_URL", "http://localhost:4000")
LITELLM_KEY = os.environ.get("LITELLM_API_KEY", "sk-litellm-charlie-2026")
VISION_MODEL = os.environ.get("VISION_MODEL", "doubao-1-5-vision-pro-32k")
CHAT_MODEL = os.environ.get("CHAT_MODEL", "openai-compatible/glm-5.1")
TEMPLATES_DIR = Path.home() / "templates"
OUTPUT_DIR = Path.home() / "output" / "docs"

# ── Document type schemas ───────────────────────────────────
DOC_SCHEMAS = {
    "customs_declaration": {
        "name": "报关单",
        "fields": [
            "申报单位",
            "收发货人",
            "贸易方式",
            "征免性质",
            "运输方式",
            "装货港",
            "目的港",
            "件数",
            "包装种类",
            "毛重(kg)",
            "净重(kg)",
            "商品编码(HS)",
            "商品名称",
            "规格型号",
            "数量",
            "单价",
            "总价",
            "币制",
            "原产国",
            "合同协议号",
        ],
    },
    "quotation": {
        "name": "报价单",
        "fields": [
            "客户名称",
            "报价日期",
            "有效期",
            "商品名称",
            "规格型号",
            "数量",
            "单位",
            "单价",
            "总价",
            "币种",
            "付款方式",
            "交货期",
            "贸易条款(Incoterms)",
            "包装方式",
            "备注",
        ],
    },
    "commercial_invoice": {
        "name": "商业发票",
        "fields": [
            "发票编号",
            "开票日期",
            "卖方名称",
            "买方名称",
            "商品描述",
            "HS编码",
            "数量",
            "单价",
            "总价",
            "币种",
            "贸易条款",
            "付款方式",
            "起运港",
            "目的港",
        ],
    },
    "packing_list": {
        "name": "装箱单",
        "fields": [
            "装箱单编号",
            "日期",
            "商品名称",
            "规格型号",
            "数量",
            "净重(kg)",
            "毛重(kg)",
            "体积(CBM)",
            "箱号",
            "包装方式",
            "总件数",
            "总净重",
            "总毛重",
            "总体积",
            "唛头",
        ],
    },
}

EXTRACT_PROMPT = """你是一个专业的外贸文档数据提取AI。请从图片中提取所有可识别的结构化信息。

请按以下JSON格式返回（只返回JSON，不要其他内容）：
{
  "doc_type": "报关单|报价单|商业发票|装箱单|其他",
  "doc_date": "YYYY-MM-DD 或 null",
  "entities": {
    "company_names": ["识别到的公司名称"],
    "person_names": ["识别到的人名"],
    "product_names": ["识别到的商品名称"]
  },
  "fields": {
    "字段名": "值",
    ...
  },
  "line_items": [
    {"商品名称": "", "HS编码": "", "数量": "", "单价": "", "总价": ""}
  ],
  "raw_text": "图片中所有可读文字",
  "confidence": 0.0-1.0
}"""

ENRICH_PROMPT = """你是外贸文档补全专家。根据已有数据补全缺失字段。

已有数据：
{existing_data}

缺失字段：{missing_fields}

已知实体信息（来自CRM）：
{entity_info}

请返回JSON格式的补全建议（只返回有把握的字段）：
{{"字段名": "建议值", ...}}"""


# ── Core Functions ──────────────────────────────────────────


def extract(file_path: str) -> dict:
    """Extract structured data from document image/PDF using doubao vision OCR."""
    path = Path(file_path)
    if not path.exists():
        return {"error": f"文件不存在: {file_path}"}

    # Read and encode file
    with open(path, "rb") as f:
        img_b64 = base64.b64encode(f.read()).decode()

    # Determine mime type
    ext = path.suffix.lower()
    mime = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".pdf": "application/pdf",
        ".webp": "image/webp",
    }.get(ext, "image/jpeg")

    try:
        with httpx.Client(timeout=60) as client:
            resp = client.post(
                f"{LITELLM_URL}/v1/chat/completions",
                headers={"Authorization": f"Bearer {LITELLM_KEY}"},
                json={
                    "model": VISION_MODEL,
                    "messages": [
                        {
                            "role": "user",
                            "content": [
                                {"type": "text", "text": EXTRACT_PROMPT},
                                {
                                    "type": "image_url",
                                    "image_url": {
                                        "url": f"data:{mime};base64,{img_b64}"
                                    },
                                },
                            ],
                        }
                    ],
                    "max_tokens": 2000,
                    "temperature": 0.1,
                },
            )
            resp.raise_for_status()
            text = resp.json()["choices"][0]["message"]["content"]
        # Parse JSON from response
        return _parse_json_response(text)
    except Exception as e:
        return {"error": f"OCR提取失败: {e}", "raw": str(e)}


def enrich(data: dict, entity_info: dict | None = None) -> dict:
    """Enrich extracted data using Context Graph and LLM."""
    if not entity_info:
        # Try to match entities from extracted data
        entity_info = _auto_match_entities(data)

    # Find missing fields
    doc_type = data.get("doc_type", "未知")
    schema_key = _guess_schema_key(doc_type)
    if schema_key and schema_key in DOC_SCHEMAS:
        expected = DOC_SCHEMAS[schema_key]["fields"]
        existing = data.get("fields", {})
        missing = [f for f in expected if f not in existing or not existing[f]]
    else:
        missing = []

    if not missing:
        return {**data, "enrichment": "无缺失字段"}

    # Ask LLM to suggest completions
    prompt = ENRICH_PROMPT.format(
        existing_data=json.dumps(data, ensure_ascii=False, indent=2),
        missing_fields=", ".join(missing),
        entity_info=json.dumps(entity_info, ensure_ascii=False, indent=2),
    )

    try:
        with httpx.Client(timeout=30) as client:
            resp = client.post(
                f"{LITELLM_URL}/v1/chat/completions",
                headers={"Authorization": f"Bearer {LITELLM_KEY}"},
                json={
                    "model": CHAT_MODEL,
                    "messages": [{"role": "user", "content": prompt}],
                    "max_tokens": 1000,
                    "temperature": 0.2,
                },
            )
            resp.raise_for_status()
            suggestions_text = resp.json()["choices"][0]["message"]["content"]
            suggestions = _parse_json_response(suggestions_text)

        # Merge suggestions into fields
        enriched_fields = {**data.get("fields", {}), **suggestions}
        return {**data, "fields": enriched_fields, "enrichment": suggestions}
    except Exception as e:
        return {**data, "enrichment_error": str(e)}


def fill_form(template_name: str, data: dict, output_name: str | None = None) -> str:
    """Fill a Word/Excel template with structured data."""
    template_path = TEMPLATES_DIR / template_name
    if not template_path.exists():
        return f"模板不存在: {template_path}"

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ext = template_path.suffix.lower()

    fields = data.get("fields", {})
    # Flatten line items into numbered fields
    line_items = data.get("line_items", [])
    for i, item in enumerate(line_items, 1):
        for k, v in item.items():
            fields[f"line_{i}_{k}"] = v

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_name = output_name or f"{template_path.stem}_{timestamp}{ext}"
    out_path = OUTPUT_DIR / out_name

    try:
        if ext == ".docx":
            return _fill_docx(template_path, out_path, fields)
        elif ext in (".xlsx", ".xls"):
            return _fill_xlsx(template_path, out_path, fields)
        else:
            return f"不支持的模板格式: {ext}"
    except Exception as e:
        return f"填充失败: {e}"


def validate(data: dict) -> dict:
    """Validate extracted data for completeness and correctness."""
    doc_type = data.get("doc_type", "未知")
    issues = []
    fields = data.get("fields", {})

    # Required fields check
    schema_key = _guess_schema_key(doc_type)
    if schema_key in DOC_SCHEMAS:
        for f in DOC_SCHEMAS[schema_key]["fields"][
            :6
        ]:  # First 6 are typically required
            if f not in fields or not fields[f]:
                issues.append({"field": f, "issue": "必填字段缺失"})

    # HS code format check
    for f_name, f_val in fields.items():
        if "HS" in f_name or "商品编码" in f_name:
            hs = str(f_val).replace(".", "").replace(" ", "")
            if hs and (not hs.isdigit() or len(hs) not in (8, 10)):
                issues.append({"field": f_name, "issue": f"HS编码格式异常: {f_val}"})

    # Numeric fields check
    for f_name in ["数量", "单价", "总价", "毛重(kg)", "净重(kg)", "体积(CBM)"]:
        if f_name in fields and fields[f_name]:
            try:
                float(str(fields[f_name]).replace(",", ""))
            except ValueError:
                issues.append({"field": f_name, "issue": "数值格式异常"})

    confidence = data.get("confidence", 0)
    return {
        "valid": len(issues) == 0,
        "issues": issues,
        "confidence": confidence,
        "doc_type": doc_type,
        "fields_count": len(fields),
        "line_items_count": len(data.get("line_items", [])),
    }


# ── Internal helpers ────────────────────────────────────────


def _parse_json_response(text: str) -> dict:
    """Extract JSON from LLM response (may have markdown fences)."""
    text = text.strip()
    if text.startswith("```"):
        lines = text.split("\n")
        text = "\n".join(lines[1:-1] if lines[-1].strip() == "```" else lines[1:])
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        # Try to find JSON object in text
        start = text.find("{")
        end = text.rfind("}") + 1
        if start >= 0 and end > start:
            try:
                return json.loads(text[start:end])
            except json.JSONDecodeError:
                pass
        return {"raw_text": text, "parse_error": True}


def _guess_schema_key(doc_type: str) -> str | None:
    mapping = {
        "报关单": "customs_declaration",
        "customs": "customs_declaration",
        "报价单": "quotation",
        "quotation": "quotation",
        "quote": "quotation",
        "商业发票": "commercial_invoice",
        "invoice": "commercial_invoice",
        "装箱单": "packing_list",
        "packing": "packing_list",
    }
    return mapping.get(doc_type)


def _auto_match_entities(data: dict) -> dict:
    """Auto-match extracted entity names to Context Graph."""
    try:
        sys.path.insert(0, str(Path.home() / "agi"))
        from context_graph import entity_match, company_full_profile

        info = {"persons": [], "companies": []}
        entities = data.get("entities", {})

        for name in entities.get("company_names", []):
            m = entity_match(name)
            if m and m.get("type") == "company":
                profile = company_full_profile(m["id"])
                info["companies"].append(profile)

        for name in entities.get("person_names", []):
            m = entity_match(name)
            if m and m.get("type") == "person":
                info["persons"].append(m)

        return info
    except Exception:
        return {}


def _fill_docx(template_path: Path, out_path: Path, fields: dict) -> str:
    """Fill .docx template using docxtpl (jinja2-based)."""
    from docxtpl import DocxTemplate

    doc = DocxTemplate(str(template_path))
    doc.render(fields)
    doc.save(str(out_path))
    return str(out_path)


def _fill_xlsx(template_path: Path, out_path: Path, fields: dict) -> str:
    """Fill .xlsx template using openpyxl with {{placeholder}} pattern."""
    from openpyxl import load_workbook
    import re

    wb = load_workbook(str(template_path))
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Replace {{field_name}} patterns
                    new_val = cell.value
                    for k, v in fields.items():
                        new_val = new_val.replace("{{" + k + "}}", str(v))
                    if new_val != cell.value:
                        cell.value = new_val
    wb.save(str(out_path))
    return str(out_path)


# ── CLI ─────────────────────────────────────────────────────


def main():
    p = argparse.ArgumentParser(description="Document Intelligence Pipeline")
    sub = p.add_subparsers(dest="cmd")

    ext = sub.add_parser("extract", help="OCR extract from file")
    ext.add_argument("file", help="Image/PDF file path")

    enr = sub.add_parser("enrich", help="Enrich extracted JSON")
    enr.add_argument("json_file", help="Extracted JSON file")

    fill = sub.add_parser("fill", help="Fill template with data")
    fill.add_argument("--template", required=True, help="Template file name")
    fill.add_argument("--data", required=True, help="JSON data file")
    fill.add_argument("--output", default=None, help="Output file name")

    val = sub.add_parser("validate", help="Validate extracted data")
    val.add_argument("json_file", help="Extracted JSON file")

    pipe = sub.add_parser("pipeline", help="Full pipeline: extract → enrich → validate")
    pipe.add_argument("file", help="Input file")
    pipe.add_argument("--template", default=None, help="Template to fill (optional)")

    args = p.parse_args()
    if not args.cmd:
        p.print_help()
        sys.exit(1)

    if args.cmd == "extract":
        result = extract(args.file)
        print(json.dumps(result, ensure_ascii=False, indent=2))

    elif args.cmd == "enrich":
        data = json.loads(Path(args.json_file).read_text())
        result = enrich(data)
        print(json.dumps(result, ensure_ascii=False, indent=2))

    elif args.cmd == "fill":
        data = json.loads(Path(args.data).read_text())
        result = fill_form(args.template, data, args.output)
        print(result)

    elif args.cmd == "validate":
        data = json.loads(Path(args.json_file).read_text())
        result = validate(data)
        print(json.dumps(result, ensure_ascii=False, indent=2))

    elif args.cmd == "pipeline":
        print("[1/3] 提取中...")
        data = extract(args.file)
        if "error" in data:
            print(json.dumps(data, ensure_ascii=False, indent=2))
            sys.exit(1)
        print(f"  → 文档类型: {data.get('doc_type', '未知')}")

        print("[2/3] 补全中...")
        data = enrich(data)
        missing = len([f for f in data.get("enrichment", {}) if data["enrichment"][f]])
        print(f"  → 补全字段: {missing}")

        print("[3/3] 验证中...")
        v = validate(data)
        print(f"  → 有效: {v['valid']}, 问题: {len(v['issues'])}")

        if args.template:
            out = fill_form(args.template, data)
            print(f"  → 输出: {out}")

        # Save intermediate JSON
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        json_out = OUTPUT_DIR / f"extract_{ts}.json"
        json_out.write_text(json.dumps(data, ensure_ascii=False, indent=2))
        print(f"\n📎 JSON → {json_out}")


if __name__ == "__main__":
    main()
